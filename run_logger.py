"""
Run logging module for the Legacy Code Documentation Agent.
Tracks processing history, token usage, and costs using JSONL for reliability.
Also maintains an Excel file for easy viewing and sharing.
"""

import os
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Log files
JSONL_FILE = Path("run_history.jsonl")
EXCEL_FILE = Path("run_history.xlsx")

# Excel styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2874A6", end_color="2874A6", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Column configuration
COLUMNS = [
    ("Timestamp", 20),
    ("Filename", 25),
    ("Language", 12),
    ("Model", 18),
    ("Input Tokens", 14),
    ("Output Tokens", 14),
    ("Total Tokens", 14),
    ("Cost (USD)", 12),
    ("Output Path", 45),
    ("PDF Path", 45),
    ("Success", 10),
    ("Error", 30)
]


def create_excel_file() -> Workbook:
    """Create a new Excel workbook with headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Run History"

    # Add headers
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    return wb


def append_to_excel(entry: dict) -> None:
    """Append a single entry to the Excel file."""
    # Create file if it doesn't exist
    if not EXCEL_FILE.exists():
        wb = create_excel_file()
    else:
        try:
            wb = load_workbook(EXCEL_FILE)
        except Exception:
            # If file is corrupted, create new one
            wb = create_excel_file()

    ws = wb.active

    # Find next empty row
    next_row = ws.max_row + 1

    # Format timestamp for readability
    timestamp = entry.get("timestamp", "")
    if timestamp:
        try:
            dt = datetime.fromisoformat(timestamp)
            timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass

    # Prepare row data
    row_data = [
        timestamp,
        entry.get("filename", ""),
        entry.get("language", ""),
        entry.get("model", ""),
        entry.get("input_tokens", 0),
        entry.get("output_tokens", 0),
        entry.get("total_tokens", 0),
        entry.get("cost_usd", 0),
        entry.get("output_path", ""),
        entry.get("pdf_path", ""),
        "Yes" if entry.get("success", True) else "No",
        entry.get("error", "") or ""
    ]

    # Write row with styling
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER

        # Format cost as currency
        if col_idx == 8 and isinstance(value, (int, float)):
            cell.number_format = "$#,##0.0000"

        # Format token columns with comma separators
        if col_idx in [5, 6, 7] and isinstance(value, (int, float)):
            cell.number_format = "#,##0"

        # Color code success/failure
        if col_idx == 11:
            if value == "Yes":
                cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    wb.save(EXCEL_FILE)


def load_history() -> list:
    """Load existing run history from JSONL file."""
    if not JSONL_FILE.exists():
        return []

    history = []
    try:
        with open(JSONL_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    try:
                        history.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
        return history
    except IOError:
        return []


def log_run(
        filename: str,
        language: str,
        model: str,
        input_tokens: int,
        output_tokens: int,
        cost: float,
        output_path: str,
        pdf_path: str = None,
        success: bool = True,
        error: str = None
) -> dict:
    """
    Log a single run to both JSONL and Excel files.
    """
    entry = {
        "timestamp": datetime.now().isoformat(),
        "filename": filename,
        "language": language,
        "model": model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
        "cost_usd": round(cost, 6),
        "output_path": output_path,
        "pdf_path": pdf_path,
        "success": success,
        "error": error
    }

    # Append to JSONL (primary, reliable storage)
    with open(JSONL_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry) + "\n")

    # Append to Excel (for easy viewing)
    try:
        append_to_excel(entry)
    except Exception as e:
        print(f"   ⚠️  Excel logging failed: {e}")

    return entry


def get_summary() -> dict:
    """Get a summary of all runs."""
    history = load_history()

    if not history:
        return {
            "total_runs": 0,
            "successful_runs": 0,
            "failed_runs": 0,
            "total_input_tokens": 0,
            "total_output_tokens": 0,
            "total_cost_usd": 0.0,
            "by_language": {},
            "by_model": {}
        }

    summary = {
        "total_runs": len(history),
        "successful_runs": sum(1 for h in history if h.get("success", True)),
        "failed_runs": sum(1 for h in history if not h.get("success", True)),
        "total_input_tokens": sum(h.get("input_tokens", 0) for h in history),
        "total_output_tokens": sum(h.get("output_tokens", 0) for h in history),
        "total_cost_usd": round(sum(h.get("cost_usd", 0) for h in history), 4),
        "by_language": {},
        "by_model": {},
        "first_run": history[0]["timestamp"] if history else None,
        "last_run": history[-1]["timestamp"] if history else None
    }

    # Group by language
    for entry in history:
        lang = entry.get("language", "Unknown")
        if lang not in summary["by_language"]:
            summary["by_language"][lang] = {"count": 0, "cost": 0.0}
        summary["by_language"][lang]["count"] += 1
        summary["by_language"][lang]["cost"] += entry.get("cost_usd", 0)

    # Group by model
    for entry in history:
        model = entry.get("model", "Unknown")
        if model not in summary["by_model"]:
            summary["by_model"][model] = {"count": 0, "cost": 0.0}
        summary["by_model"][model]["count"] += 1
        summary["by_model"][model]["cost"] += entry.get("cost_usd", 0)

    return summary


def print_summary() -> None:
    """Print a formatted summary to the console."""
    summary = get_summary()

    print("\n" + "=" * 60)
    print("   RUN HISTORY SUMMARY")
    print("=" * 60)

    if summary["total_runs"] == 0:
        print("\n   No runs recorded yet.")
        print("=" * 60 + "\n")
        return

    print(f"\n   Total runs:      {summary['total_runs']}")
    print(f"   Successful:      {summary['successful_runs']}")
    print(f"   Failed:          {summary['failed_runs']}")
    print("-" * 60)
    print(f"   Total tokens:    {summary['total_input_tokens']:,} in / {summary['total_output_tokens']:,} out")
    print(f"   Total cost:      ${summary['total_cost_usd']:.4f}")
    print("-" * 60)

    print("\n   By Language:")
    for lang, data in summary["by_language"].items():
        print(f"      {lang:<15}: {data['count']:<4} files (${data['cost']:.4f})")

    print("\n   By Model:")
    for model, data in summary["by_model"].items():
        print(f"      {model:<15}: {data['count']:<4} runs (${data['cost']:.4f})")

    print("-" * 60)
    print(f"   📊 Excel report: {EXCEL_FILE.absolute()}")
    print("=" * 60 + "\n")